from openpyxl import Workbook, worksheet
from openpyxl.styles import Font
from typing import Any, List
from services.ProductService import get_productname_by_id


def insert_data_to_row(data: List[Any], ws: worksheet, row):
    a = 'A'
    for d in data:
        c = ws[a+str(row)]
        c.value = d
        a = chr(ord(a)+1)


async def generate_xlfile(data, filename):
    wb = Workbook()
    ws = wb.create_sheet()
    i = 1
    for j in range(len(data)):
        i = i+1
        if (j > 0 and data[j]["superproduct_id"] == data[j-1]["superproduct_id"]):
            insert_data_to_row(data[j][:-1], ws, i)
        else:
            product_name = await get_productname_by_id(data[j]["superproduct_id"])
            insert_data_to_row(
                [product_name], ws, i)
            i = i+1
            insert_data_to_row(data[j][:-1], ws, i)
    wb.save(filename)
